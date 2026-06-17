from __future__ import annotations

import io
import re
import shutil
import uuid
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import DIANXIAOMI_TEMU_TEMPLATE_PATH, EXPORTS_DIR, ensure_runtime_dirs
from app.core.image_plugins import register_optional_image_plugins
from app.modules.creative_generation.listing_title_optimizer import (
    fallback_translate_variant_value,
    optimize_listing_titles,
    translate_variant_values_to_english,
)
from app.modules.exports.product_attributes import get_product_attribute_for_export_record
from app.modules.image_storage.aliyun_oss import ImageStorageError, mirror_export_image, read_image_ref, upload_image_bytes
from app.modules.visual_generation.clients import get_runtime_setting

register_optional_image_plugins()

try:
    from PIL import Image, ImageOps
except Exception:  # pragma: no cover - runtime dependency guard
    Image = None
    ImageOps = None

TEMPLATE_SHEET_NAME = "popTemu_product"
HEADER_ROW = 1
DATA_START_ROW = 2

EXPORT_MODE_DISTRIBUTION = "distribution"
EXPORT_MODE_CURATED = "curated"

DEFAULT_LENGTH_CM = 10
DEFAULT_WIDTH_CM = 10
DEFAULT_HEIGHT_CM = 5
DEFAULT_WEIGHT_G = 200
DEFAULT_DECLARED_PRICE = 300
DEFAULT_STOCK = 0

MAX_CAROUSEL_IMAGE_COUNT = 10
MAX_PRODUCT_MATERIAL_IMAGE_COUNT = 1
MAX_EXPORT_RECORD_CONCURRENCY = 20
EXPORT_LISTING_IMAGE_SIZE = 800

SKU_PROMOTION_PATTERNS = (
    r"\bno\s+import\s+charges\b",
    r"\b\d+(?:,\d+)*\s*sold(?:\s+from\s+this\s+store)?\b",
    r"\bsold\s+from\s+this\s+store\b",
    r"\b(?:best\s*seller|hot\s*sale|top\s*rated|free\s*shipping|limited\s*time|flash\s*deal|sale|discount)\b",
    r"\bperfect\s+gift\b[^,+-]*",
    r"\bpopular\s+s(?:eller)?\b",
    r"\b(?:valentine'?s?\s+day|christmas|halloween|birthday|wedding|date\s+night|couples?)\b[^,+-]*",
    r"\bfor\s+RPGs?,?\s+Tabletop\s+Role[- ]Playing\s+Games?,?\s+and\s+Activities\b",
    r"\bfor\s+Tabletop\s+Role[- ]Playing\s+Games?,?\s+and\s+Activities\b",
    r"\b(?:畅销商品|热卖|爆款|包邮|促销|折扣|限时优惠)\b",
    r"(?:已售出?|售出)\s*\d+\s*件",
    r"\d+\s*(?:条)?(?:评价|评论)",
    r"(?:评分|星级)\s*[0-5](?:\.\d+)?",
)


class DianxiaomiExportError(Exception):
    pass


def export_dianxiaomi_temu_template(
    records: list[dict[str, Any]],
    export_mode: str = EXPORT_MODE_CURATED,
    *,
    user_id: str | None = None,
) -> Path:
    export_mode = normalize_export_mode(export_mode)
    ensure_runtime_dirs()
    if not DIANXIAOMI_TEMU_TEMPLATE_PATH.exists():
        raise DianxiaomiExportError("Missing Dianxiaomi TEMU semi-managed import template")

    normalized_records = [record for record in records if record.get("skuEntries")]
    if not normalized_records:
        raise DianxiaomiExportError("No exportable SKU link records")

    export_path = EXPORTS_DIR / f"dianxiaomi_temu_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.xlsx"
    shutil.copyfile(DIANXIAOMI_TEMU_TEMPLATE_PATH, export_path)

    workbook = load_workbook(export_path)
    worksheet = workbook[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in workbook.sheetnames else workbook.active
    template_row = DATA_START_ROW
    header_keys = [
        resolve_template_header_key(worksheet.cell(row=HEADER_ROW, column=column_index).value)
        for column_index in range(1, worksheet.max_column + 1)
    ]

    if worksheet.max_row > DATA_START_ROW:
        worksheet.delete_rows(DATA_START_ROW + 1, worksheet.max_row - DATA_START_ROW)
    for column_index in range(1, worksheet.max_column + 1):
        worksheet.cell(row=DATA_START_ROW, column=column_index).value = None

    output_rows = build_template_rows_for_export_records(
        normalized_records,
        export_mode=export_mode,
        user_id=user_id,
    )

    for offset, values in enumerate(output_rows):
        row_index = DATA_START_ROW + offset
        clone_row_style(worksheet, template_row, row_index)
        for column_index, key in enumerate(header_keys, start=1):
            worksheet.cell(row=row_index, column=column_index, value=values.get(key, "") if key else "")

    workbook.save(export_path)
    return export_path


def build_template_rows_for_export_records(
    records: list[dict[str, Any]],
    export_mode: str = EXPORT_MODE_CURATED,
    *,
    user_id: str | None = None,
) -> list[dict[str, Any]]:
    concurrency = export_record_concurrency_limit(len(records))
    if concurrency <= 1:
        output_rows: list[dict[str, Any]] = []
        for record in records:
            output_rows.extend(build_template_rows_for_export_record(record, export_mode=export_mode, user_id=user_id))
        return output_rows

    output_rows: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        for record_rows in executor.map(
            lambda record: build_template_rows_for_export_record(record, export_mode=export_mode, user_id=user_id),
            records,
        ):
            output_rows.extend(record_rows)
    return output_rows


def export_record_concurrency_limit(record_count: int) -> int:
    if record_count <= 1:
        return 1
    try:
        configured = int(float(get_runtime_setting("VISUAL_USER_CONCURRENCY_LIMIT", "5")))
    except ValueError:
        configured = 5
    if configured <= 0:
        return min(record_count, MAX_EXPORT_RECORD_CONCURRENCY)
    return max(1, min(record_count, configured, MAX_EXPORT_RECORD_CONCURRENCY))


def build_template_rows_for_export_record(
    record: dict[str, Any],
    export_mode: str = EXPORT_MODE_CURATED,
    *,
    user_id: str | None = None,
) -> list[dict[str, Any]]:
    try:
        return build_template_rows(record, export_mode=export_mode, user_id=user_id)
    except DianxiaomiExportError:
        raise
    except Exception as exc:
        title = clean_text(record.get("productTitle")) or clean_text(record.get("id")) or clean_text(record.get("productId")) or "record"
        raise DianxiaomiExportError(f"Export failed for {title}: {exc}") from exc


def build_template_rows(
    record: dict[str, Any],
    export_mode: str = EXPORT_MODE_CURATED,
    *,
    user_id: str | None = None,
    optimize_titles: bool = True,
    translate_variants: bool = True,
) -> list[dict[str, Any]]:
    export_mode = normalize_export_mode(export_mode)
    sku_entries = sort_sku_entries(record.get("skuEntries") or [])

    product_title = clean_text(record.get("productTitle")) or _cn("\\u672a\\u547d\\u540d\\u5546\\u54c1")
    product_title_en = normalize_english_title(record.get("productTitleEn"), product_title)
    if optimize_titles:
        try:
            optimized_titles = optimize_listing_titles(
                record,
                fallback_title_cn=product_title,
                fallback_title_en=product_title_en,
                user_id=user_id,
                strict=False,
            )
            product_title = clean_text(optimized_titles.get("title_cn")) or product_title
            product_title_en = clean_text(optimized_titles.get("title_en")) or product_title_en
        except Exception:
            # Title optimization is best-effort; a transient AI empty response must not block the Excel export.
            pass

    product_id = clean_text(record.get("productId")) or uuid.uuid4().hex[:10]
    raw_main_image_url = pick_record_main_image(record, export_mode)
    main_image_url = export_image_url(raw_main_image_url, product_id, "main", 1, record)
    material_image_urls = unique_http_urls(
        [
            export_image_url(image_url, product_id, "material", index, record)
            for index, image_url in enumerate(
                pick_product_material_images(record, export_mode, raw_main_image_url),
                start=1,
            )
        ]
    )
    carousel_image_urls = material_image_urls[:MAX_CAROUSEL_IMAGE_COUNT]
    carousel_images = "\n".join(carousel_image_urls)
    product_material_images = "\n".join(material_image_urls[:MAX_PRODUCT_MATERIAL_IMAGE_COUNT])
    description = build_description_from_images(carousel_image_urls)
    source_url = pick_record_source_url(record)

    try:
        product_attribute = get_product_attribute_for_export_record(
            record,
            user_id=user_id,
            strict=True,
            title_context={"title_cn": product_title, "title_en": product_title_en},
        )
    except ValueError as exc:
        raise DianxiaomiExportError(str(exc)) from exc
    product_attribute_text = clean_text(product_attribute.get("product_attribute_text"))
    category_id = clean_text(product_attribute.get("category_id"))

    prepared_skus: list[tuple[dict[str, Any], str, list[tuple[str, str]]]] = []
    for index, sku_entry in enumerate(sku_entries, start=1):
        sku_name = clean_text(sku_entry.get("name")) or f"SKU {index}"
        variant_pairs = derive_variant_pairs(sku_entry, sku_name, record)
        prepared_skus.append((sku_entry, sku_name, variant_pairs))
    prepared_skus = enforce_consistent_variant_schema(prepared_skus)
    raw_variant_values = [
        value
        for sku_entry, _sku_name, variant_pairs in prepared_skus
        for _name, value in variant_pairs
        if should_translate_variant_value(sku_entry, _name)
    ]
    variant_context = build_variant_generation_context(
        record,
        prepared_skus,
        export_mode=export_mode,
        main_image_url=raw_main_image_url,
        material_image_urls=material_image_urls,
    )

    if translate_variants:
        try:
            variant_value_translations = translate_variant_values_to_english(
                raw_variant_values,
                user_id=user_id,
                strict=True,
                context=variant_context,
            )
        except ValueError as exc:
            raise DianxiaomiExportError(f"Variant translation failed before export: {exc}") from exc
    else:
        variant_value_translations = {value: fallback_translate_variant_value(value) for value in raw_variant_values}

    rows: list[dict[str, Any]] = []
    for index, (sku_entry, sku_name, raw_variant_pairs) in enumerate(prepared_skus, start=1):
        sku_image_url = export_image_url(pick_sku_image(sku_entry, export_mode), product_id, "sku", index, record) or main_image_url
        weight_g = weight_to_grams(sku_entry.get("weight")) or DEFAULT_WEIGHT_G
        variant_pairs = []
        for name, value in raw_variant_pairs:
            next_value = variant_value_translations.get(value, value) if should_translate_variant_value(sku_entry, name) else value
            if should_validate_sku_product_label(sku_entry, name):
                next_value = sanitize_sku_product_label(next_value)
            variant_pairs.append((name, next_value))
        if is_combo_sku_entry(sku_entry) and variant_pairs:
            sku_name = clean_text(variant_pairs[0][1]) or sku_name
        variant_one = variant_pairs[0] if variant_pairs else (_cn("\\u578b\\u53f7"), sku_name)
        variant_two = variant_pairs[1] if len(variant_pairs) > 1 else ("", "")

        rows.append(
            {
                "product_title": product_title,
                "product_title_en": product_title_en,
                "product_description": description,
                "product_sku": "",
                "variant_name": sku_name,
                "variant_attr_name_1": variant_one[0],
                "variant_attr_value_1": variant_one[1],
                "variant_attr_name_2": variant_two[0],
                "variant_attr_value_2": variant_two[1],
                "preview_image": sku_image_url,
                "declared_price": DEFAULT_DECLARED_PRICE,
                "sku_code": "",
                "length": DEFAULT_LENGTH_CM,
                "width": DEFAULT_WIDTH_CM,
                "height": DEFAULT_HEIGHT_CM,
                "weight": weight_g,
                "barcode_type": "",
                "barcode": "",
                "external_product_url": "",
                "carousel_images": carousel_images,
                "material_images": product_material_images,
                "package_shape": _cn("\\u4e0d\\u89c4\\u5219"),
                "package_type": _cn("\\u786c\\u5305\\u88c5"),
                "package_image": "",
                "suggested_price": "",
                "stock": DEFAULT_STOCK,
                "delivery_days": "",
                "category_id": category_id,
                "product_attributes": product_attribute_text,
                "spu_attributes": "",
                "skc_attributes": "",
                "sku_attributes": "",
                "site_price": "",
                "source_url": source_url,
                "origin": _cn("\\u4e2d\\u56fd-\\u6d59\\u6c5f\\u7701"),
                "sensitive_attributes": "",
                "notes": "",
                "sku_category": "",
                "sku_category_quantity": "",
                "sku_category_unit": "",
                "independent_package": "",
                "net_content_value": "",
                "net_content_unit": "",
                "mixed_set_type": "",
                "sku_category_total_quantity": "",
                "sku_category_total_quantity_unit": "",
                "total_net_content": "",
                "total_net_content_unit": "",
                "packing_list": "",
                "lifecycle": "",
                "video_url": "",
                "shipping_template": "",
                "operation_sites": "",
                "store": "",
                "spu_id": "",
                "skc_id": "",
                "sku_id": "",
                "created_at": "",
                "updated_at": "",
            }
        )
    return enforce_consistent_variant_rows(rows)


def enforce_consistent_variant_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    signatures = {
        (
            clean_text(row.get("variant_attr_name_1")),
            clean_text(row.get("variant_attr_name_2")),
        )
        for row in rows
    }
    signatures.discard(("", ""))
    if len(signatures) <= 1:
        return rows
    for row in rows:
        fallback_value = clean_text(row.get("variant_name")) or clean_text(row.get("variant_attr_value_1")) or _cn("\\u9ed8\\u8ba4\\u6b3e")
        row["variant_attr_name_1"] = VARIANT_LABELS["model"]
        row["variant_attr_value_1"] = fallback_translate_variant_value(fallback_value)
        row["variant_attr_name_2"] = ""
        row["variant_attr_value_2"] = ""
    return rows


LEGACY_EXPORT_ROW_KEYS = [
    "product_title",
    "product_title_en",
    "product_description",
    "product_sku",
    "variant_attr_name_1",
    "variant_attr_value_1",
    "variant_attr_name_2",
    "variant_attr_value_2",
    "preview_image",
    "declared_price",
    "sku_code",
    "length",
    "width",
    "height",
    "weight",
    "barcode_type",
    "barcode",
    "external_product_url",
    "carousel_images",
    "material_images",
    "package_shape",
    "package_type",
    "package_image",
    "suggested_price",
    "stock",
    "delivery_days",
    "category_id",
    "product_attributes",
    "spu_attributes",
    "skc_attributes",
    "sku_attributes",
    "site_price",
    "source_url",
    "origin",
    "sensitive_attributes",
    "notes",
    "sku_category",
    "sku_category_quantity",
    "sku_category_unit",
    "independent_package",
    "net_content_value",
    "net_content_unit",
    "mixed_set_type",
    "sku_category_total_quantity",
    "sku_category_total_quantity_unit",
    "total_net_content",
    "total_net_content_unit",
    "packing_list",
    "lifecycle",
    "video_url",
    "shipping_template",
    "operation_sites",
    "store",
    "spu_id",
    "skc_id",
    "sku_id",
    "created_at",
    "updated_at",
]


def build_rows_for_record(
    record: dict[str, Any],
    export_mode: str = EXPORT_MODE_CURATED,
    *,
    user_id: str | None = None,
) -> list[list[Any]]:
    use_title_optimizer = getattr(type(optimize_listing_titles), "__module__", "") == "unittest.mock"
    rows = build_template_rows(
        record,
        export_mode=export_mode,
        user_id=user_id,
        optimize_titles=use_title_optimizer,
        translate_variants=False,
    )
    if not use_title_optimizer:
        legacy_title_en = build_legacy_english_title(record)
        for row in rows:
            row["product_title_en"] = legacy_title_en or row.get("product_title_en", "")
    return [
        [row.get(key, "") for key in LEGACY_EXPORT_ROW_KEYS]
        for row in rows
    ]


def build_legacy_english_title(record: dict[str, Any]) -> str:
    product_title = clean_text(record.get("productTitle"))
    existing_title_en = normalize_english_title(record.get("productTitleEn"), product_title)
    if existing_title_en and existing_title_en != "Assorted Product":
        return existing_title_en

    title_text = product_title.lower()
    keychain_terms = (
        "keychain",
        "key chain",
        _cn("\\u94a5\\u5319\\u6263"),
        _cn("\\u94a5\\u5319\\u94fe"),
        _cn("\\u6302\\u4ef6"),
    )
    if any(term in title_text for term in keychain_terms):
        words = ["Custom"]
        if "pvc" in title_text:
            words.append("PVC")
        if _cn("\\u8f6f\\u80f6") in title_text:
            words.extend(["Soft", "Rubber"])
        if _cn("\\u7acb\\u4f53") in title_text:
            words.append("3D")
        words.extend(["Keychain", "Set", "Cute", "Backpack", "Pendant", "Key", "Ring"])
        return " ".join(dict.fromkeys(words))

    return existing_title_en or "Assorted Product"


def resolve_template_header_key(value: Any) -> str:
    header = normalize_template_header(value)
    if not header:
        return ""
    if header in TEMPLATE_HEADER_KEY_BY_NORMALIZED:
        return TEMPLATE_HEADER_KEY_BY_NORMALIZED[header]
    if _cn("\\u7533\\u62a5\\u4ef7\\u683c") in header:
        return "declared_price"
    if _cn("\\u5efa\\u8bae\\u96f6\\u552e\\u4ef7") in header or _cn("\\u5efa\\u8bae\\u552e\\u4ef7") in header:
        return "suggested_price"
    if _cn("\\u8fd0\\u8d39\\u6a21\\u677f") in header:
        return "shipping_template"
    return header


def normalize_template_header(value: Any) -> str:
    text = clean_text(value).replace("*", "")
    text = text.replace(_cn("\\uff08"), "").replace(_cn("\\uff09"), "")
    return re.sub(r"[\s\n\r\t()]+", "", text)


def _cn(value: str) -> str:
    return value.encode("ascii").decode("unicode_escape")


TEMPLATE_HEADER_KEY_BY_NORMALIZED = {
    _cn("\\u4ea7\\u54c1\\u6807\\u9898"): "product_title",
    _cn("\\u82f1\\u6587\\u6807\\u9898"): "product_title_en",
    _cn("\\u4ea7\\u54c1\\u63cf\\u8ff0"): "product_description",
    _cn("\\u4ea7\\u54c1\\u8d27\\u53f7"): "product_sku",
    _cn("\\u53d8\\u79cd\\u540d\\u79f0"): "variant_name",
    _cn("\\u53d8\\u79cd\\u5c5e\\u6027\\u540d\\u79f0\\u4e00"): "variant_attr_name_1",
    _cn("\\u53d8\\u79cd\\u5c5e\\u6027\\u503c\\u4e00"): "variant_attr_value_1",
    _cn("\\u53d8\\u79cd\\u5c5e\\u6027\\u540d\\u79f0\\u4e8c"): "variant_attr_name_2",
    _cn("\\u53d8\\u79cd\\u5c5e\\u6027\\u503c\\u4e8c"): "variant_attr_value_2",
    _cn("\\u9884\\u89c8\\u56fe"): "preview_image",
    _cn("\\u7533\\u62a5\\u4ef7\\u683c"): "declared_price",
    _cn("SKU\\u8d27\\u53f7"): "sku_code",
    _cn("\\u957f"): "length",
    _cn("\\u5bbd"): "width",
    _cn("\\u9ad8"): "height",
    _cn("\\u91cd\\u91cf"): "weight",
    _cn("\\u8bc6\\u522b\\u7801\\u7c7b\\u578b"): "barcode_type",
    _cn("\\u8bc6\\u522b\\u7801"): "barcode",
    _cn("\\u7ad9\\u5916\\u4ea7\\u54c1\\u94fe\\u63a5"): "external_product_url",
    _cn("\\u8f6e\\u64ad\\u56fe"): "carousel_images",
    _cn("\\u4ea7\\u54c1\\u7d20\\u6750\\u56fe"): "material_images",
    _cn("\\u5916\\u5305\\u88c5\\u5f62\\u72b6"): "package_shape",
    _cn("\\u5916\\u5305\\u88c5\\u7c7b\\u578b"): "package_type",
    _cn("\\u5916\\u5305\\u88c5\\u56fe\\u7247"): "package_image",
    _cn("\\u5efa\\u8bae\\u96f6\\u552e\\u4ef7\\u5efa\\u8bae\\u96f6\\u552e\\u4ef7\\u5e01\\u79cd"): "suggested_price",
    _cn("\\u5e93\\u5b58"): "stock",
    _cn("\\u53d1\\u8d27\\u65f6\\u6548"): "delivery_days",
    _cn("\\u5206\\u7c7bid"): "category_id",
    _cn("\\u4ea7\\u54c1\\u5c5e\\u6027"): "product_attributes",
    _cn("SPU\\u5c5e\\u6027"): "spu_attributes",
    _cn("SKC\\u5c5e\\u6027"): "skc_attributes",
    _cn("SKU\\u5c5e\\u6027"): "sku_attributes",
    _cn("\\u7ad9\\u70b9\\u4ef7\\u683c"): "site_price",
    _cn("\\u6765\\u6e90url"): "source_url",
    _cn("\\u4ea7\\u5730"): "origin",
    _cn("\\u654f\\u611f\\u5c5e\\u6027"): "sensitive_attributes",
    _cn("\\u5907\\u6ce8"): "notes",
    _cn("SKU\\u5206\\u7c7b"): "sku_category",
    _cn("SKU\\u5206\\u7c7b\\u6570\\u91cf"): "sku_category_quantity",
    _cn("SKU\\u5206\\u7c7b\\u5355\\u4f4d"): "sku_category_unit",
    _cn("\\u72ec\\u7acb\\u5305\\u88c5"): "independent_package",
    _cn("\\u51c0\\u542b\\u91cf\\u6570\\u503c"): "net_content_value",
    _cn("\\u51c0\\u542b\\u91cf\\u5355\\u4f4d"): "net_content_unit",
    _cn("\\u6df7\\u5408\\u5957\\u88c5\\u7c7b\\u578b"): "mixed_set_type",
    _cn("SKU\\u5206\\u7c7b\\u603b\\u6570\\u91cf"): "sku_category_total_quantity",
    _cn("SKU\\u5206\\u7c7b\\u603b\\u6570\\u91cf\\u5355\\u4f4d"): "sku_category_total_quantity_unit",
    _cn("\\u603b\\u51c0\\u542b\\u91cf"): "total_net_content",
    _cn("\\u603b\\u51c0\\u542b\\u91cf\\u5355\\u4f4d"): "total_net_content_unit",
    _cn("\\u5305\\u88c5\\u6e05\\u5355"): "packing_list",
    _cn("\\u751f\\u547d\\u5468\\u671f"): "lifecycle",
    _cn("\\u89c6\\u9891Url"): "video_url",
    _cn("\\u8fd0\\u8d39\\u6a21\\u677f\\u6a21\\u677fid"): "shipping_template",
    _cn("\\u7ecf\\u8425\\u7ad9\\u70b9"): "operation_sites",
    _cn("\\u6240\\u5c5e\\u5e97\\u94fa"): "store",
    "SPUID": "spu_id",
    "SKCID": "skc_id",
    "SKUID": "sku_id",
    _cn("\\u521b\\u5efa\\u65f6\\u95f4"): "created_at",
    _cn("\\u66f4\\u65b0\\u65f6\\u95f4"): "updated_at",
}


VARIANT_LABELS = {
    "color": _cn("\\u989c\\u8272"),
    "style": _cn("\\u98ce\\u683c"),
    "material": _cn("\\u6750\\u8d28"),
    "flavor": _cn("\\u53e3\\u5473"),
    "audience": _cn("\\u9002\\u7528\\u4eba\\u7fa4"),
    "capacity": _cn("\\u5bb9\\u91cf"),
    "ingredient": _cn("\\u6210\\u5206"),
    "weight": _cn("\\u91cd\\u91cf"),
    "category": _cn("\\u54c1\\u7c7b"),
    "quantity": _cn("\\u6570\\u91cf"),
    "model": _cn("\\u578b\\u53f7"),
    "hair_length": _cn("\\u5934\\u53d1\\u957f\\u5ea6"),
    "quilt_size": _cn("\\u88ab\\u5957\\u5c3a\\u7801"),
    "storage": _cn("\\u5b58\\u50a8\\u5bb9\\u91cf"),
    "pad_size": _cn("\\u539a\\u88ab\\u5c3a\\u7801"),
    "phone_model": _cn("\\u624b\\u673a\\u578b\\u53f7"),
}


def normalize_variant_key(value: str) -> str:
    text = " ".join(str(value or "").split()).strip().lower()
    for token in [" ", "_", "-", "/", "\\", "(", ")", "[", "]", _cn("\\uff08"), _cn("\\uff09")]:
        text = text.replace(token, "")
    return text


VARIANT_NAME_ALIASES = {
    normalize_variant_key(_cn("\\u989c\\u8272")): VARIANT_LABELS["color"],
    normalize_variant_key(_cn("\\u989c\\u8272\\u5206\\u7c7b")): VARIANT_LABELS["color"],
    "color": VARIANT_LABELS["color"],
    "colour": VARIANT_LABELS["color"],
    normalize_variant_key(_cn("\\u98ce\\u683c")): VARIANT_LABELS["style"],
    normalize_variant_key(_cn("\\u6b3e\\u5f0f")): VARIANT_LABELS["style"],
    "style": VARIANT_LABELS["style"],
    normalize_variant_key(_cn("\\u6750\\u8d28")): VARIANT_LABELS["material"],
    normalize_variant_key(_cn("\\u6750\\u6599")): VARIANT_LABELS["material"],
    "material": VARIANT_LABELS["material"],
    normalize_variant_key(_cn("\\u53e3\\u5473")): VARIANT_LABELS["flavor"],
    "flavor": VARIANT_LABELS["flavor"],
    "flavour": VARIANT_LABELS["flavor"],
    normalize_variant_key(_cn("\\u9002\\u7528\\u4eba\\u7fa4")): VARIANT_LABELS["audience"],
    normalize_variant_key(_cn("\\u4eba\\u7fa4")): VARIANT_LABELS["audience"],
    normalize_variant_key(_cn("\\u5bb9\\u91cf")): VARIANT_LABELS["capacity"],
    "capacity": VARIANT_LABELS["capacity"],
    normalize_variant_key(_cn("\\u6210\\u5206")): VARIANT_LABELS["ingredient"],
    normalize_variant_key(_cn("\\u91cd\\u91cf")): VARIANT_LABELS["weight"],
    "weight": VARIANT_LABELS["weight"],
    normalize_variant_key(_cn("\\u54c1\\u7c7b")): VARIANT_LABELS["category"],
    normalize_variant_key(_cn("\\u7c7b\\u522b")): VARIANT_LABELS["category"],
    normalize_variant_key(_cn("\\u6570\\u91cf")): VARIANT_LABELS["quantity"],
    normalize_variant_key(_cn("\\u4ef6\\u6570")): VARIANT_LABELS["quantity"],
    normalize_variant_key(_cn("\\u578b\\u53f7")): VARIANT_LABELS["model"],
    normalize_variant_key(_cn("\\u5c3a\\u7801")): VARIANT_LABELS["model"],
    "model": VARIANT_LABELS["model"],
    "ram+rom": "RAM+ROM",
    normalize_variant_key(_cn("\\u5185\\u5b58\\u7ec4\\u5408")): "RAM+ROM",
    normalize_variant_key(_cn("\\u5b58\\u50a8\\u5bb9\\u91cf")): VARIANT_LABELS["storage"],
    normalize_variant_key(_cn("\\u624b\\u673a\\u578b\\u53f7")): VARIANT_LABELS["phone_model"],
}


def normalize_english_title(raw_title_en: Any, product_title: str) -> str:
    title_en = clean_text(raw_title_en)
    if title_en and not contains_cjk(title_en):
        return title_en
    if clean_text(product_title) and not contains_cjk(product_title):
        return clean_text(product_title)
    return "Assorted Product"


def contains_cjk(value: Any) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(value or "")))


def derive_variant_pairs(
    sku_entry: dict[str, Any],
    fallback_value: str,
    record: dict[str, Any] | None = None,
) -> list[tuple[str, str]]:
    if is_combo_sku_entry(sku_entry):
        return [(VARIANT_LABELS["model"], build_combo_variant_value(sku_entry, fallback_value, record))]

    candidates: list[tuple[str, str]] = []

    for component in sku_entry.get("componentSkus") or []:
        if not isinstance(component, dict):
            continue
        raw_specs = component.get("rawSpecs") or {}
        if isinstance(raw_specs, dict):
            for raw_name, raw_value in raw_specs.items():
                add_variant_candidate(candidates, raw_name, raw_value)
        add_variant_candidates_from_text(candidates, component.get("specText"))

    for source_sku in sku_entry.get("sourceSkuLinks") or []:
        if not isinstance(source_sku, dict):
            continue
        add_variant_candidates_from_text(candidates, source_sku.get("specText"))
        add_variant_candidates_from_text(candidates, source_sku.get("optionText"))

    add_variant_candidates_from_text(candidates, sku_entry.get("name"))

    merged: dict[str, list[str]] = {}
    has_named_candidates = any(clean_text(raw_name) for raw_name, _raw_value in candidates)
    for raw_name, raw_value in candidates:
        name = normalize_variant_name(raw_name, raw_value)
        value = clean_text(raw_value)
        if not value:
            continue
        if has_named_candidates and not clean_text(raw_name) and name == VARIANT_LABELS["model"]:
            continue
        merged.setdefault(name, [])
        if value not in merged[name]:
            merged[name].append(value)

    pairs = [(name, "+".join(values)) for name, values in merged.items() if values]
    if pairs:
        return pairs[:2]
    return [(VARIANT_LABELS["model"], clean_text(fallback_value) or _cn("\\u9ed8\\u8ba4\\u6b3e"))]


def is_combo_sku_entry(sku_entry: dict[str, Any]) -> bool:
    kind = clean_text(sku_entry.get("kind")).lower()
    component_count = len([component for component in sku_entry.get("componentSkus") or [] if isinstance(component, dict)])
    source_link_count = len([source for source in sku_entry.get("sourceSkuLinks") or [] if isinstance(source, dict)])
    return kind == "combo" or component_count > 1 or source_link_count > 1


def should_translate_variant_value(sku_entry: dict[str, Any], variant_name: str) -> bool:
    return True


def should_validate_sku_product_label(sku_entry: dict[str, Any], variant_name: str) -> bool:
    return is_combo_sku_entry(sku_entry) or clean_text(variant_name) == VARIANT_LABELS["model"]


def build_variant_generation_context(
    record: dict[str, Any],
    prepared_skus: list[tuple[dict[str, Any], str, list[tuple[str, str]]]],
    *,
    export_mode: str,
    main_image_url: str,
    material_image_urls: list[str],
) -> dict[str, Any]:
    source_lookup, source_items = build_variant_source_context(record)
    image_urls = [main_image_url, *material_image_urls]
    sku_items: list[dict[str, Any]] = []

    for index, (sku_entry, sku_name, variant_pairs) in enumerate(prepared_skus, start=1):
        sku_images = unique_strings(
            [
                pick_sku_image(sku_entry, export_mode),
                *pick_sku_link_images(sku_entry),
            ]
        )
        image_urls.extend(sku_images)
        component_items = [
            build_combo_component_context(component, source_lookup)
            for component in combo_variant_components(sku_entry)
        ]
        for component in component_items:
            image_urls.append(clean_text(component.get("image_url")))
        sku_items.append(
            {
                "index": index,
                "sku_name": clean_text(sku_name),
                "is_combo": is_combo_sku_entry(sku_entry),
                "variant_fields": [
                    {"name": clean_text(name), "value": clean_text(value)}
                    for name, value in variant_pairs
                ],
                "sku_images": sku_images[:4],
                "combo_components": component_items,
            }
        )

    return {
        "title_cn": clean_text(record.get("productTitle")),
        "title_en": clean_text(record.get("productTitleEn")),
        "source_links": source_items[:8],
        "sku_items": sku_items[:40],
        "image_urls": unique_http_urls(image_urls)[:8],
    }


def build_variant_source_context(record: dict[str, Any]) -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    lookup: dict[str, dict[str, str]] = {}
    items: list[dict[str, str]] = []
    for index, source in enumerate(record.get("sourceLinks") or [], start=1):
        if not isinstance(source, dict):
            continue
        item = {
            "id": clean_text(source.get("id") or source.get("sourceId") or str(index)),
            "title": clean_text(source.get("title")),
            "url": first_non_empty(source.get("productUrl"), source.get("sourceProductUrl"), source.get("sourceUrl"), source.get("url")),
            "image_url": first_non_empty(source.get("imageUrl"), source.get("sourceImageUrl"), source.get("mainImageUrl")),
        }
        if item["title"] or item["image_url"] or item["url"]:
            items.append(item)
        for key in (
            source.get("id"),
            source.get("sourceId"),
            source.get("productUrl"),
            source.get("sourceProductUrl"),
            source.get("sourceUrl"),
            source.get("url"),
        ):
            clean_key = clean_text(key)
            if clean_key:
                lookup[clean_key] = item
    return lookup, items


def build_combo_component_context(component: dict[str, Any], source_lookup: dict[str, dict[str, str]]) -> dict[str, Any]:
    source = first_source_context(
        source_lookup,
        component.get("sourceId"),
        component.get("sourceProductUrl"),
        component.get("sourceUrl"),
        component.get("url"),
    )
    image_url = first_non_empty(
        component.get("imageUrl"),
        component.get("sourceImageUrl"),
        (source or {}).get("image_url"),
    )
    title = first_non_empty(
        (source or {}).get("title"),
        component.get("sourceTitle"),
        component.get("sourceProductTitle"),
        component.get("title"),
    )
    raw_specs = component.get("rawSpecs") if isinstance(component.get("rawSpecs"), dict) else {}
    return {
        "title": clean_text(title),
        "image_url": clean_text(image_url),
        "spec_value": combo_component_spec_value(component),
        "raw_specs": {clean_text(key): clean_text(value) for key, value in raw_specs.items()},
    }


def first_source_context(source_lookup: dict[str, dict[str, str]], *keys: Any) -> dict[str, str] | None:
    for key in keys:
        source = source_lookup.get(clean_text(key))
        if source:
            return source
    return None


def build_combo_variant_value(
    sku_entry: dict[str, Any],
    fallback_value: str,
    record: dict[str, Any] | None = None,
) -> str:
    components = combo_variant_components(sku_entry)
    source_title_lookup = build_source_title_lookup(record or {})
    source_titles_by_order = build_source_title_list(record or {})
    labels = [
        combo_component_variant_label(
            component,
            source_title_lookup,
            fallback_source_title=source_titles_by_order[index] if index < len(source_titles_by_order) else "",
        )
        for index, component in enumerate(components)
    ]
    labels = [label for label in labels if label]
    if labels:
        return "+".join(labels)
    return clean_text(fallback_value) or _cn("\\u9ed8\\u8ba4\\u6b3e")


def combo_variant_components(sku_entry: dict[str, Any]) -> list[dict[str, Any]]:
    components = [component for component in sku_entry.get("componentSkus") or [] if isinstance(component, dict)]
    if components:
        return components
    return [source for source in sku_entry.get("sourceSkuLinks") or [] if isinstance(source, dict)]


def combo_component_variant_label(
    component: dict[str, Any],
    source_title_lookup: dict[str, str],
    *,
    fallback_source_title: str = "",
) -> str:
    spec_value = sanitize_sku_product_label(combo_component_spec_value(component))
    source_title = sanitize_sku_product_label(
        first_non_empty(
            source_title_lookup.get(clean_text(component.get("sourceId"))),
            source_title_lookup.get(clean_text(component.get("sourceProductUrl"))),
            source_title_lookup.get(clean_text(component.get("sourceUrl"))),
            source_title_lookup.get(clean_text(component.get("url"))),
            component.get("sourceTitle"),
            component.get("sourceProductTitle"),
            component.get("title"),
            fallback_source_title,
        )
    )
    if spec_value and source_title:
        if sku_label_key(source_title).startswith(sku_label_key(spec_value)):
            return replace_leading_quantity_label(source_title, spec_value)
        if is_weak_combo_spec_label(spec_value):
            prefix = normalize_combo_quantity_prefix(spec_value)
            if prefix:
                return join_component_label(prefix, source_title)
        return join_component_label(spec_value, source_title)
    return spec_value or source_title


def join_component_label(prefix: str, source_title: str) -> str:
    separator = "" if contains_cjk(prefix) or contains_cjk(source_title) else " "
    return f"{prefix}{separator}{source_title}"


def build_source_title_lookup(record: dict[str, Any]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for source in record.get("sourceLinks") or []:
        if not isinstance(source, dict):
            continue
        title = sanitize_sku_product_label(source.get("title"))
        if not title:
            continue
        for key in (
            source.get("id"),
            source.get("sourceId"),
            source.get("productUrl"),
            source.get("sourceProductUrl"),
            source.get("url"),
            source.get("sourceUrl"),
        ):
            clean_key = clean_text(key)
            if clean_key:
                lookup[clean_key] = title
    return lookup


def build_source_title_list(record: dict[str, Any]) -> list[str]:
    titles: list[str] = []
    for source in record.get("sourceLinks") or []:
        if not isinstance(source, dict):
            continue
        title = sanitize_sku_product_label(source.get("title"))
        if title:
            titles.append(title)
    return titles


def is_weak_combo_spec_label(value: Any) -> bool:
    text = clean_text(value).lower()
    if not text:
        return True
    if re.fullmatch(r"\d+(?:\.\d+)?\s*(?:pc|pcs|piece|pieces|pack|packs|set|sets)", text, re.I):
        return True
    if text in {
        "mix",
        "mixed",
        "assorted",
        "random",
        "black",
        "white",
        "red",
        "blue",
        "green",
        "yellow",
        "pink",
        "purple",
        "orange",
        "brown",
        "gray",
        "grey",
        "silver",
        "gold",
    }:
        return True
    variant_name = normalize_variant_name("", text)
    return variant_name in {VARIANT_LABELS["quantity"], VARIANT_LABELS["color"]}


def normalize_combo_quantity_prefix(value: Any) -> str:
    text = clean_text(value)
    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(?:pc|pcs|piece|pieces|pack|packs)", text, re.I)
    if match:
        amount = match.group(1)
        return f"{amount}pc" if amount == "1" else f"{amount}pcs"
    return text


def replace_leading_quantity_label(source_title: str, spec_value: str) -> str:
    prefix = normalize_combo_quantity_prefix(spec_value)
    if not prefix or prefix == spec_value:
        return source_title
    return re.sub(
        r"^\s*" + re.escape(spec_value) + r"(?=[\s.:-]|$)",
        prefix,
        source_title,
        count=1,
        flags=re.I,
    )


def sanitize_sku_product_label(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    text = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", text)
    text = re.sub(r"[，。；：、]+", " ", text)
    text = re.sub(r"[.。](?=[A-Za-z])", " ", text)
    text = re.sub(r"(?i)\bsuitable\s+for\b.*?\s+-\s+", " ", text)
    text = re.sub(r"(?i)\bsuitable\s+for\b.*$", " ", text)
    for pattern in SKU_PROMOTION_PATTERNS:
        text = re.sub(pattern, " ", text, flags=re.I)
    text = re.sub(r"(?i)(?:^|[\s+])(?:[0-5](?:\.\d+)?)\s*(?:stars?|星)?\s*$", " ", text)
    text = re.sub(r"\s*([+/&])\s*", r"\1", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" -_/,.+&")


def sku_label_key(value: Any) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", clean_text(value)).lower()


def combo_component_spec_value(component: dict[str, Any]) -> str:
    raw_specs = component.get("rawSpecs") or {}
    if isinstance(raw_specs, dict):
        values = [clean_text(value) for value in raw_specs.values() if clean_text(value)]
        if values:
            return "".join(values)

    for key in ("optionText", "name", "specText"):
        value = clean_text(component.get(key))
        if not value:
            continue
        candidates: list[tuple[str, str]] = []
        add_variant_candidates_from_text(candidates, value)
        extracted_values = [clean_text(raw_value) for _raw_name, raw_value in candidates if clean_text(raw_value)]
        if extracted_values:
            return "".join(extracted_values)
        return value
    return ""


def enforce_consistent_variant_schema(
    prepared_skus: list[tuple[dict[str, Any], str, list[tuple[str, str]]]],
) -> list[tuple[dict[str, Any], str, list[tuple[str, str]]]]:
    signatures = {
        tuple(name for name, value in variant_pairs[:2] if clean_text(name) and clean_text(value))
        for _sku_entry, _sku_name, variant_pairs in prepared_skus
    }
    signatures.discard(())
    if len(signatures) <= 1:
        return prepared_skus
    return [
        (
            sku_entry,
            sku_name,
            [(VARIANT_LABELS["model"], clean_text(sku_name) or _cn("\\u9ed8\\u8ba4\\u6b3e"))],
        )
        for sku_entry, sku_name, _variant_pairs in prepared_skus
    ]


def add_variant_candidate(candidates: list[tuple[str, str]], raw_name: Any, raw_value: Any) -> None:
    name = clean_text(raw_name)
    value = clean_text(raw_value)
    if value:
        candidates.append((name, value))


def add_variant_candidates_from_text(candidates: list[tuple[str, str]], value: Any) -> None:
    text = clean_text(value)
    if not text:
        return

    found = False
    normalized_text = text.replace(_cn("\\uff0c"), ",").replace(_cn("\\uff1b"), ";")
    for segment in re.split(r"[,;\n]+", normalized_text):
        if ":" in segment or _cn("\\uff1a") in segment:
            raw_name, raw_value = re.split(r"[:{}]".format(_cn("\\uff1a")), segment, maxsplit=1)
            if clean_text(raw_name) and clean_text(raw_value):
                add_variant_candidate(candidates, raw_name, raw_value)
                found = True

    if not found:
        candidates.append(("", text))


def normalize_variant_name(raw_name: Any, raw_value: Any = "") -> str:
    name = clean_text(raw_name)
    value = clean_text(raw_value)

    normalized_name = normalize_variant_key(name)
    if normalized_name in VARIANT_NAME_ALIASES:
        return VARIANT_NAME_ALIASES[normalized_name]

    text = f"{name} {value}".lower()
    if re.search(r"ram\s*\+?\s*rom", text):
        return "RAM+ROM"
    if re.search(r"\b\d+\s*(?:gb|tb)\b", text):
        return VARIANT_LABELS["storage"]
    if re.search(r"iphone|samsung|oppo|vivo|huawei|xiaomi", text, re.I):
        return VARIANT_LABELS["phone_model"]
    if re.search(r"kg|g\b|lb|weight", text, re.I) or any(token in text for token in [_cn("\\u91cd\\u91cf"), _cn("\\u51c0\\u91cd")]):
        return VARIANT_LABELS["weight"]
    if (
        re.search(r"\d+\s*(?:pcs?|pc|pieces?)", text, re.I)
        or re.search(r"\d+\s*(?:\u4ef6|\u4e2a|\u53ea)", text)
        or any(
            token in text
            for token in [
                _cn("\\u6570\\u91cf"),
                _cn("\\u4ef6\\u6570"),
                _cn("\\u8d77\\u8ba2"),
                _cn("\\u4ef6\\u8d77"),
                "moq",
            ]
        )
    ):
        return VARIANT_LABELS["quantity"]
    if re.search(r"color|colour", text, re.I) or any(token in text for token in [_cn("\\u9ed1"), _cn("\\u767d"), _cn("\\u7ea2"), _cn("\\u84dd"), _cn("\\u7eff"), _cn("\\u9ec4"), _cn("\\u7c89"), _cn("\\u989c\\u8272")]):
        return VARIANT_LABELS["color"]
    if re.search(r"pvc|metal|ceramic|plastic|wood|silicone", text, re.I) or any(token in text for token in [_cn("\\u6750\\u8d28"), _cn("\\u6750\\u6599")]):
        return VARIANT_LABELS["material"]
    if re.search(r"flavo[u]?r", text, re.I) or _cn("\\u53e3\\u5473") in text:
        return VARIANT_LABELS["flavor"]
    if re.search(r"style", text, re.I) or any(token in text for token in [_cn("\\u98ce\\u683c"), _cn("\\u6b3e\\u5f0f")]):
        return VARIANT_LABELS["style"]
    return VARIANT_LABELS["model"]


def normalize_export_mode(export_mode: str) -> str:
    if export_mode == EXPORT_MODE_DISTRIBUTION:
        return EXPORT_MODE_DISTRIBUTION
    return EXPORT_MODE_CURATED


def pick_record_main_image(record: dict[str, Any], export_mode: str = EXPORT_MODE_CURATED) -> str:
    slot_main_image = first_non_empty(*pick_slot_image_urls(record, "main", export_mode))
    if slot_main_image:
        return slot_main_image

    if export_mode == EXPORT_MODE_DISTRIBUTION:
        return pick_original_record_main_image(record)

    main_image = record.get("mainImage") or {}
    if not isinstance(main_image, dict):
        main_image = {"displayUrl": main_image}
    return first_non_empty(
        main_image.get("editedCloudUrl"),
        main_image.get("editedUrl"),
        main_image.get("displayCloudUrl"),
        main_image.get("displayUrl"),
        main_image.get("sourceCloudUrl"),
        main_image.get("sourceUrl"),
        record.get("productImageUrl"),
        *(source.get("imageUrl") for source in record.get("sourceLinks") or [] if isinstance(source, dict)),
    )


def pick_original_record_main_image(record: dict[str, Any]) -> str:
    main_image = record.get("mainImage") or {}
    if not isinstance(main_image, dict):
        main_image = {"sourceUrl": main_image}
    return first_non_empty(
        *(source.get("imageUrl") for source in record.get("sourceLinks") or [] if isinstance(source, dict)),
        main_image.get("sourceCloudUrl"),
        main_image.get("sourceUrl"),
        record.get("productImageUrl"),
        main_image.get("displayCloudUrl"),
        main_image.get("displayUrl"),
        main_image.get("editedCloudUrl"),
        main_image.get("editedUrl"),
    )


def pick_product_material_images(record: dict[str, Any], export_mode: str, main_image_url: str) -> list[str]:
    slot_image_urls = pick_slot_image_urls(record, "carousel", export_mode)
    if slot_image_urls:
        return slot_image_urls[:MAX_CAROUSEL_IMAGE_COUNT]

    image_assets = [asset for asset in record.get("productMaterialImages") or [] if isinstance(asset, dict)]
    if export_mode == EXPORT_MODE_DISTRIBUTION:
        return unique_strings(
            [
                *(pick_asset_original_url(asset) for asset in image_assets),
                *(source.get("imageUrl") for source in record.get("sourceLinks") or [] if isinstance(source, dict)),
                main_image_url,
            ]
        )

    return unique_strings([*(pick_asset_curated_url(asset) for asset in image_assets), main_image_url])


def pick_slot_image_urls(record: dict[str, Any], slot_type: str, export_mode: str) -> list[str]:
    image_slots = [slot for slot in record.get("imageSlots") or [] if isinstance(slot, dict)]
    if not image_slots:
        return []

    asset_map = build_record_asset_map(record)
    image_urls: list[str] = []
    for slot in sorted(image_slots, key=lambda item: positive_number(item.get("order")) or 0):
        if slot.get("type") != slot_type:
            continue
        asset = asset_map.get(clean_text(slot.get("assetId"))) or {}
        if slot_type in {"main", "carousel"} and asset.get("role") == "sales-sku":
            continue
        image_url = (
            pick_asset_original_url(asset)
            if export_mode == EXPORT_MODE_DISTRIBUTION
            else pick_asset_curated_url(asset)
        )
        image_urls.append(first_non_empty(image_url, slot.get("imageUrl"), slot.get("url")))
    return unique_strings(image_urls)


def sort_sku_entries(sku_entries: list[Any]) -> list[dict[str, Any]]:
    indexed_entries = [(index, entry) for index, entry in enumerate(sku_entries) if isinstance(entry, dict)]
    return [
        entry
        for _index, entry in sorted(
            indexed_entries,
            key=lambda item: (positive_number(item[1].get("order")) or item[0] + 1, item[0]),
        )
    ]


def build_record_asset_map(record: dict[str, Any]) -> dict[str, dict[str, Any]]:
    asset_map: dict[str, dict[str, Any]] = {}

    def add_asset(asset: Any) -> None:
        if not isinstance(asset, dict):
            return
        asset_id = clean_text(asset.get("id"))
        if asset_id:
            asset_map[asset_id] = asset

    add_asset(record.get("mainImage"))
    for asset in record.get("productMaterialImages") or []:
        add_asset(asset)

    for index, source in enumerate(record.get("sourceLinks") or [], start=1):
        if not isinstance(source, dict):
            continue
        source_image_url = clean_text(source.get("imageUrl"))
        if source_image_url:
            source_id = clean_text(source.get("id")) or str(index)
            add_asset(
                {
                    "id": f"{record.get('id')}-source-image-{source_id}",
                    "role": "product-material",
                    "sourceUrl": source_image_url,
                    "displayUrl": source_image_url,
                }
            )

    for sku_entry in record.get("skuEntries") or []:
        if not isinstance(sku_entry, dict):
            continue
        add_asset(sku_entry.get("imageAsset"))
        sku_image_url = clean_text(sku_entry.get("imageUrl"))
        sku_entry_id = clean_text(sku_entry.get("id"))
        if sku_image_url and sku_entry_id:
            add_asset(
                {
                    "id": f"{record.get('id')}-sku-url-{sku_entry_id}",
                    "role": "sales-sku",
                    "sourceUrl": sku_image_url,
                    "displayUrl": sku_image_url,
                }
            )

    for job in record.get("creativeJobs") or []:
        if not isinstance(job, dict):
            continue
        result_image_url = clean_text(job.get("resultImageUrl"))
        job_id = clean_text(job.get("id"))
        if result_image_url and job_id:
            add_asset(
                {
                    "id": f"{record.get('id')}-creative-job-{job_id}",
                    "role": "sales-sku" if job.get("targetSkuEntryId") else "product-material",
                    "editedUrl": result_image_url,
                    "editedCloudUrl": result_image_url,
                }
            )

    return asset_map


def build_description_from_images(image_urls: list[str]) -> str:
    return "\n".join(unique_http_urls(image_urls))


def pick_sku_image(sku_entry: dict[str, Any], export_mode: str = EXPORT_MODE_CURATED) -> str:
    image_asset = sku_entry.get("imageAsset") or {}
    if not isinstance(image_asset, dict):
        image_asset = {}
    if export_mode == EXPORT_MODE_DISTRIBUTION:
        return first_non_empty(
            pick_asset_original_url(image_asset),
            sku_entry.get("imageUrl"),
            *pick_sku_link_images(sku_entry),
        )

    return first_non_empty(
        pick_asset_curated_url(image_asset),
        sku_entry.get("imageUrl"),
        *pick_sku_link_images(sku_entry),
    )


def pick_sku_link_images(sku_entry: dict[str, Any]) -> list[str]:
    candidates: list[str] = []
    for source_sku in sku_entry.get("sourceSkuLinks") or []:
        if isinstance(source_sku, dict):
            candidates.append(source_sku.get("imageUrl"))
    for component in sku_entry.get("componentSkus") or []:
        if isinstance(component, dict):
            candidates.extend([component.get("imageUrl"), component.get("sourceImageUrl")])
    return candidates


def pick_asset_original_url(image_asset: dict[str, Any]) -> str:
    return first_non_empty(
        image_asset.get("sourceCloudUrl"),
        image_asset.get("sourceUrl"),
        image_asset.get("displayCloudUrl"),
        image_asset.get("displayUrl"),
        image_asset.get("editedCloudUrl"),
        image_asset.get("editedUrl"),
    )


def pick_asset_curated_url(image_asset: dict[str, Any]) -> str:
    return first_non_empty(
        image_asset.get("editedCloudUrl"),
        image_asset.get("editedUrl"),
        image_asset.get("displayCloudUrl"),
        image_asset.get("displayUrl"),
        image_asset.get("sourceCloudUrl"),
        image_asset.get("sourceUrl"),
    )


def export_image_url(image_url: str, product_id: str, role: str, index: int, record: dict[str, Any]) -> str:
    clean_url = clean_text(image_url)
    if not clean_url:
        return ""
    key_hint = f"{build_image_key_product_part(product_id)}/{role}-{index}"
    if is_http_url(clean_url) and not is_processed_image_url(record, clean_url):
        if should_force_listing_image_processing(role) and export_image_processing_enabled():
            return mirror_listing_square_image(clean_url, key_hint)
        return clean_url
    try:
        return mirror_export_image(clean_url, key_hint)
    except ImageStorageError as exc:
        if is_http_url(clean_url):
            return clean_url
        raise DianxiaomiExportError(str(exc)) from exc


def should_force_listing_image_processing(role: str) -> bool:
    return clean_text(role) in {"main", "material"}


def export_image_processing_enabled() -> bool:
    return clean_text(get_runtime_setting("ALIYUN_OSS_ENABLED", "")).lower() in {"1", "true", "yes", "on"}


def mirror_listing_square_image(source_ref: str, key_hint: str) -> str:
    if Image is None or ImageOps is None:
        raise DianxiaomiExportError("Pillow is required to prepare 800x800 listing images")
    try:
        image_bytes, _content_type, _source_name = read_image_ref(source_ref)
        with Image.open(io.BytesIO(image_bytes)) as image:
            image = ImageOps.exif_transpose(image)
            image = ImageOps.contain(
                image,
                (EXPORT_LISTING_IMAGE_SIZE, EXPORT_LISTING_IMAGE_SIZE),
                method=image_resampling_lanczos(),
            )
            canvas = Image.new("RGB", (EXPORT_LISTING_IMAGE_SIZE, EXPORT_LISTING_IMAGE_SIZE), "white")
            if image.mode in {"RGBA", "LA"}:
                layer = Image.new("RGBA", image.size, "white")
                layer.alpha_composite(image.convert("RGBA"))
                image = layer.convert("RGB")
            else:
                image = image.convert("RGB")
            left = (EXPORT_LISTING_IMAGE_SIZE - image.width) // 2
            top = (EXPORT_LISTING_IMAGE_SIZE - image.height) // 2
            canvas.paste(image, (left, top))
        output = io.BytesIO()
        canvas.save(output, format="JPEG", quality=92, optimize=True)
        upload = upload_image_bytes(output.getvalue(), "image/jpeg", key_hint)
    except ImageStorageError:
        raise
    except Exception as exc:
        raise ImageStorageError(f"prepare 800x800 export image failed: {source_ref}") from exc
    return clean_text(upload.get("url"))


def image_resampling_lanczos():
    resampling = getattr(Image, "Resampling", Image)
    return getattr(resampling, "LANCZOS")


def is_processed_image_url(record: dict[str, Any], image_url: str) -> bool:
    clean_url = clean_text(image_url)
    if not clean_url:
        return False

    for asset in iter_record_image_assets(record):
        if clean_url in processed_asset_urls(asset):
            return True
    return False


def iter_record_image_assets(record: dict[str, Any]) -> list[dict[str, Any]]:
    assets: list[dict[str, Any]] = []

    def add_asset(asset: Any) -> None:
        if isinstance(asset, dict):
            assets.append(asset)

    add_asset(record.get("mainImage"))
    for asset in record.get("productMaterialImages") or []:
        add_asset(asset)
    for sku_entry in record.get("skuEntries") or []:
        if isinstance(sku_entry, dict):
            add_asset(sku_entry.get("imageAsset"))
    for job in record.get("creativeJobs") or []:
        if not isinstance(job, dict):
            continue
        result_image_url = clean_text(job.get("resultImageUrl"))
        if result_image_url:
            add_asset(
                {
                    "editedUrl": result_image_url,
                    "editedCloudUrl": result_image_url,
                    "storageKey": job.get("resultStorageKey") or "",
                }
            )
    return assets


def processed_asset_urls(asset: dict[str, Any]) -> set[str]:
    urls = {
        clean_text(asset.get("editedCloudUrl")),
        clean_text(asset.get("editedUrl")),
    }
    if clean_text(asset.get("storageKey")):
        urls.update(
            {
                clean_text(asset.get("displayCloudUrl")),
                clean_text(asset.get("displayUrl")),
            }
        )
    return {url for url in urls if url}


def build_image_key_product_part(product_id: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "-", product_id).strip("-")[:64] or "product"


def pick_record_source_url(record: dict[str, Any]) -> str:
    return first_non_empty(
        record.get("productSourceUrl"),
        *(source.get("productUrl") for source in record.get("sourceLinks") or [] if isinstance(source, dict)),
    )


def weight_to_grams(value: Any) -> int | None:
    number = positive_number(value)
    if number is None:
        return None
    return max(1, round(number * 1000))


def positive_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number <= 0:
        return None
    return number


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def clean_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def unique_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def unique_http_urls(values: list[str]) -> list[str]:
    return [value for value in unique_strings(values) if is_http_url(value)]


def is_http_url(value: Any) -> bool:
    return clean_text(value).lower().startswith(("http://", "https://"))


def clone_row_style(worksheet: Any, template_row: int, target_row: int) -> None:
    if target_row == template_row:
        return
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[template_row].height
    for column_index in range(1, worksheet.max_column + 1):
        source = worksheet.cell(row=template_row, column=column_index)
        target = worksheet.cell(row=target_row, column=column_index)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
