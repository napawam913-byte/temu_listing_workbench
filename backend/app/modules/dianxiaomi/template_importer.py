from __future__ import annotations

import json
import re
import shutil
import uuid
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.core.config import UPLOADS_DIR, ensure_runtime_dirs
from app.core.database import PRODUCT_CATALOG_SCOPE_POOL_ONLY, utc_now_text
from app.modules.exports.dianxiaomi_temu import (
    DATA_START_ROW,
    HEADER_ROW,
    TEMPLATE_SHEET_NAME,
    clean_text,
    resolve_template_header_key,
)
from app.modules.link_records.postgres_store import upsert_link_list_records
from app.modules.products.postgres_store import get_pg_connection


class DianxiaomiTemplateImportError(Exception):
    pass


def import_dianxiaomi_template_file(
    file_obj: BinaryIO,
    filename: str,
    *,
    user_id: str,
) -> dict[str, Any]:
    ensure_runtime_dirs()
    batch_id = uuid.uuid4().hex
    safe_filename = Path(filename).name
    saved_path = UPLOADS_DIR / f"{batch_id}_{safe_filename}"
    with saved_path.open("wb") as target:
        shutil.copyfileobj(file_obj, target)

    rows = read_template_rows(saved_path)
    product_groups = group_template_rows(rows)
    if not product_groups:
        raise DianxiaomiTemplateImportError("未识别到店小秘标准模板中的商品数据")

    now = utc_now_text()
    product_payloads: list[dict[str, Any]] = []
    for index, group_rows in enumerate(product_groups, start=1):
        product = build_product_payload(group_rows, batch_id=batch_id, source_row_index=index, now=now)
        product_payloads.append(product)

    upsert_products_to_pool(
        product_payloads,
        batch_id=batch_id,
        source_filename=safe_filename,
        saved_path=saved_path,
        total_rows=len(rows),
        user_id=user_id,
        now=now,
    )
    link_records = [
        build_link_list_record(product, group_rows, now=now)
        for product, group_rows in zip(product_payloads, product_groups, strict=False)
    ]
    saved_records = upsert_link_list_records(link_records, user_id=user_id)

    return {
        "batch_id": batch_id,
        "source_filename": safe_filename,
        "file_type": "xlsx",
        "total_rows": len(rows),
        "imported_count": len(product_payloads),
        "failed_count": 0,
        "errors": [],
        "records": saved_records,
    }


def read_template_rows(path: Path) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise DianxiaomiTemplateImportError(f"无法读取店小秘 Excel：{exc}") from exc

    worksheet = workbook[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in workbook.sheetnames else workbook.active
    header_keys = [
        resolve_template_header_key(worksheet.cell(row=HEADER_ROW, column=column_index).value)
        for column_index in range(1, worksheet.max_column + 1)
    ]
    if not any(header_keys):
        raise DianxiaomiTemplateImportError("未识别到店小秘模板表头")

    rows: list[dict[str, Any]] = []
    carry: dict[str, Any] = {}
    carry_keys = {
        "product_title",
        "product_title_en",
        "product_description",
        "product_sku",
        "carousel_images",
        "material_images",
        "category_id",
        "product_attributes",
        "source_url",
        "external_product_url",
        "video_url",
    }
    for row_index in range(DATA_START_ROW, worksheet.max_row + 1):
        row: dict[str, Any] = {"source_row_index": row_index}
        for column_index, key in enumerate(header_keys, start=1):
            if not key:
                continue
            value = normalize_cell_value(worksheet.cell(row=row_index, column=column_index).value)
            row[key] = value

        starts_new_product = any(
            clean_text(row.get(key))
            for key in ("product_title", "product_title_en", "source_url", "external_product_url")
        )
        if starts_new_product:
            for key in carry_keys:
                if clean_text(row.get(key)):
                    carry[key] = row[key]
                else:
                    carry.pop(key, None)
        else:
            for key in carry_keys:
                if not clean_text(row.get(key)) and clean_text(carry.get(key)):
                    row[key] = carry[key]
                elif clean_text(row.get(key)):
                    carry[key] = row[key]

        if is_empty_template_row(row):
            continue
        rows.append(row)
    return rows


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def is_empty_template_row(row: dict[str, Any]) -> bool:
    keys = ("product_title", "product_title_en", "variant_name", "preview_image", "source_url")
    return not any(clean_text(row.get(key)) for key in keys)


def group_template_rows(rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []
    for row in rows:
        key = product_group_key(row)
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(row)
    return [grouped[key] for key in order]


def product_group_key(row: dict[str, Any]) -> str:
    explicit = first_non_empty(row.get("product_sku"), row.get("spu_id"), row.get("source_url"))
    if explicit:
        return stable_slug(explicit)
    title = first_non_empty(row.get("product_title"), row.get("product_title_en"))
    category_id = clean_text(row.get("category_id"))
    return stable_slug(f"{title}|{category_id}") or uuid.uuid4().hex


def build_product_payload(group_rows: list[dict[str, Any]], *, batch_id: str, source_row_index: int, now: str) -> dict[str, Any]:
    first = group_rows[0]
    title = first_non_empty(first.get("product_title"), first.get("product_title_en"), "店小秘导入商品")
    title_en = clean_text(first.get("product_title_en"))
    category_id = clean_text(first.get("category_id"))
    source_product_id = first_non_empty(first.get("product_sku"), first.get("spu_id"), product_group_key(first))
    source_url = first_non_empty(first.get("source_url"), first.get("external_product_url"))
    image_urls = collect_product_image_urls(group_rows)
    main_image = first_non_empty(image_urls[0] if image_urls else "", first.get("preview_image"))
    price = first_positive_number(
        *(row.get("site_price") for row in group_rows),
        *(row.get("suggested_price") for row in group_rows),
        *(row.get("declared_price") for row in group_rows),
    )

    return {
        "id": product_import_id(source_product_id),
        "upload_batch_id": batch_id,
        "source_row_index": source_row_index,
        "source_type": "custom",
        "source_product_id": source_product_id,
        "catalog_scope": PRODUCT_CATALOG_SCOPE_POOL_ONLY,
        "title_cn": title,
        "title_en": title_en or None,
        "title": title,
        "main_image_url": main_image or None,
        "gallery_image_urls": image_urls,
        "video_url": clean_text(first.get("video_url")) or None,
        "source_url": source_url or None,
        "category_path": f"店小秘分类/{category_id}" if category_id else "店小秘导入",
        "category_level1": "店小秘导入",
        "category_level2": category_id or None,
        "tags": ["店小秘导入"],
        "price_usd": price,
        "gmv_usd": 0,
        "weekly_sales": 0,
        "monthly_sales": 0,
        "review_count": 0,
        "listing_time": now,
        "status": "active",
        "in_product_pool": True,
        "raw_data": {
            "source": "dianxiaomi_template",
            "rows": group_rows,
        },
    }


def build_link_list_record(product: dict[str, Any], group_rows: list[dict[str, Any]], *, now: str) -> dict[str, Any]:
    record_id = f"link-entry-dxm-{product['id']}"
    source_url = clean_text(product.get("source_url"))
    source_id = f"{record_id}-source-1"
    main_image_url = clean_text(product.get("main_image_url"))
    material_urls = collect_product_image_urls(group_rows)
    source_title = clean_text(product.get("title")) or "店小秘导入货源"
    main_image_asset_id = f"{record_id}-main-image"
    material_assets = [
        {
            "id": f"{record_id}-material-image-{index}",
            "role": "product-material",
            "sourceUrl": image_url,
            "displayUrl": image_url,
            "alt": f"{source_title} 素材图 {index}",
        }
        for index, image_url in enumerate(material_urls, start=1)
    ]
    image_slots = [
        {
            "id": f"{record_id}-slot-main",
            "type": "main",
            "order": 0,
            "assetId": main_image_asset_id,
        },
        *[
            {
                "id": f"{record_id}-slot-carousel-{index}",
                "type": "carousel",
                "order": index,
                "assetId": asset["id"],
            }
            for index, asset in enumerate(material_assets[:8], start=1)
        ],
    ]
    sku_entries = [build_sku_entry(record_id, source_id, product, row, index) for index, row in enumerate(group_rows, start=1)]

    return {
        "schemaVersion": 3,
        "id": record_id,
        "createdAt": now,
        "productId": product["id"],
        "productTitle": source_title,
        "productTitleEn": clean_text(product.get("title_en")) or None,
        "category": product.get("category_path"),
        "categoryLevel1": product.get("category_level1"),
        "categoryLevel2": product.get("category_level2"),
        "categoryPath": product.get("category_path"),
        "categoryId": clean_text(group_rows[0].get("category_id")) or None,
        "dxmCategoryId": clean_text(group_rows[0].get("category_id")) or None,
        "mainImage": {
            "id": main_image_asset_id,
            "role": "product-main",
            "sourceUrl": main_image_url,
            "displayUrl": main_image_url,
            "alt": source_title,
        },
        "productMaterialImages": material_assets,
        "imageSlots": image_slots,
        "productImageGenerationCount": 8,
        "productImageUrl": main_image_url,
        "productSourceUrl": source_url,
        "sourceLinks": [
            {
                "id": source_id,
                "title": source_title,
                "productUrl": source_url,
                "shopName": "店小秘模板",
                "imageUrl": main_image_url,
            }
        ],
        "skuEntries": sku_entries,
        "componentSkuCount": len(sku_entries),
    }


def build_sku_entry(
    record_id: str,
    source_id: str,
    product: dict[str, Any],
    row: dict[str, Any],
    index: int,
) -> dict[str, Any]:
    sku_id = f"{record_id}-sku-{index}"
    sku_name = first_non_empty(row.get("variant_name"), variant_spec_text(row), f"SKU {index}")
    sku_image_url = first_non_empty(row.get("preview_image"), product.get("main_image_url"))
    sku_code = first_non_empty(row.get("sku_code"), row.get("sku_id"), f"SKU-{index}")
    spec_text = variant_spec_text(row) or sku_name
    source_title = clean_text(product.get("title")) or "店小秘导入货源"
    source_url = clean_text(product.get("source_url"))
    weight_kg = (first_positive_number(row.get("weight")) or 0) / 1000

    source_sku_link = {
        "sourceId": source_id,
        "sourceTitle": source_title,
        "sourceProductUrl": source_url,
        "sourceSkuId": sku_code,
        "sourceSkuKey": sku_code,
        "specText": spec_text,
        "optionText": sku_name,
        "imageUrl": sku_image_url,
    }
    return {
        "id": sku_id,
        "order": index,
        "kind": "single",
        "name": sku_name,
        "imageAsset": {
            "id": f"{record_id}-sku-image-{index}",
            "role": "sales-sku",
            "sourceUrl": sku_image_url,
            "displayUrl": sku_image_url,
            "alt": sku_name,
        },
        "imageUrl": sku_image_url,
        "price": first_positive_number(row.get("site_price"), row.get("suggested_price"), row.get("declared_price")),
        "weight": weight_kg or None,
        "sourceSkuLinks": [source_sku_link],
        "componentSkus": [
            {
                "name": sku_name,
                "specText": spec_text,
                "sourceId": source_id,
                "sourceSkuId": sku_code,
                "sourceSkuKey": sku_code,
                "sourceTitle": source_title,
                "sourceUrl": source_url,
                "sourceImageUrl": clean_text(product.get("main_image_url")),
                "imageUrl": sku_image_url,
                "rawSpecs": variant_raw_specs(row),
            }
        ],
    }


def upsert_products_to_pool(
    products: list[dict[str, Any]],
    *,
    batch_id: str,
    source_filename: str,
    saved_path: Path,
    total_rows: int,
    user_id: str,
    now: str,
) -> None:
    with get_pg_connection() as conn:
        conn.execute(
            """
            INSERT INTO upload_batches (
                id, source_filename, saved_path, file_type, total_rows,
                imported_count, failed_count, status, error_message, created_at, updated_at
            ) VALUES (%s, %s, %s, 'dianxiaomi-template', %s, %s, 0, 'imported', NULL, %s, %s)
            """,
            (batch_id, source_filename, str(saved_path), total_rows, len(products), now, now),
        )
        for product in products:
            existing = conn.execute(
                """
                SELECT id, created_at
                FROM products
                WHERE source_type = %s AND source_product_id = %s
                ORDER BY updated_at DESC
                LIMIT 1
                """,
                (product["source_type"], product["source_product_id"]),
            ).fetchone()
            product_id = existing["id"] if existing else product["id"]
            conn.execute(
                """
                INSERT INTO products (
                    id, upload_batch_id, source_row_index, source_type, source_product_id, catalog_scope,
                    title_cn, title_en, title, main_image_url, gallery_image_urls_json,
                    video_url, source_url, category_path, category_level1, category_level2,
                    tags_json, price_usd, gmv_usd, weekly_sales, monthly_sales,
                    review_count, listing_time, status, in_product_pool, raw_data_json, created_at, updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s
                )
                ON CONFLICT (id) DO UPDATE SET
                    upload_batch_id = EXCLUDED.upload_batch_id,
                    source_row_index = EXCLUDED.source_row_index,
                    title_cn = EXCLUDED.title_cn,
                    title_en = EXCLUDED.title_en,
                    title = EXCLUDED.title,
                    main_image_url = EXCLUDED.main_image_url,
                    gallery_image_urls_json = EXCLUDED.gallery_image_urls_json,
                    video_url = EXCLUDED.video_url,
                    source_url = EXCLUDED.source_url,
                    category_path = EXCLUDED.category_path,
                    category_level1 = EXCLUDED.category_level1,
                    category_level2 = EXCLUDED.category_level2,
                    tags_json = EXCLUDED.tags_json,
                    price_usd = EXCLUDED.price_usd,
                    gmv_usd = EXCLUDED.gmv_usd,
                    weekly_sales = EXCLUDED.weekly_sales,
                    monthly_sales = EXCLUDED.monthly_sales,
                    review_count = EXCLUDED.review_count,
                    listing_time = EXCLUDED.listing_time,
                    status = 'active',
                    in_product_pool = 1,
                    raw_data_json = EXCLUDED.raw_data_json,
                    updated_at = EXCLUDED.updated_at
                """,
                (
                    product_id,
                    batch_id,
                    product["source_row_index"],
                    product["source_type"],
                    product["source_product_id"],
                    PRODUCT_CATALOG_SCOPE_POOL_ONLY,
                    product["title_cn"],
                    product["title_en"],
                    product["title"],
                    product["main_image_url"],
                    json.dumps(product["gallery_image_urls"], ensure_ascii=False),
                    product["video_url"],
                    product["source_url"],
                    product["category_path"],
                    product["category_level1"],
                    product["category_level2"],
                    json.dumps(product["tags"], ensure_ascii=False),
                    product["price_usd"],
                    product["gmv_usd"],
                    product["weekly_sales"],
                    product["monthly_sales"],
                    product["review_count"],
                    product["listing_time"],
                    product["status"],
                    1,
                    json.dumps(product["raw_data"], ensure_ascii=False),
                    existing["created_at"] if existing else now,
                    now,
                ),
            )
            product["id"] = product_id
            conn.execute(
                """
                INSERT INTO product_pool_memberships (user_id, product_id, status, created_at, updated_at)
                VALUES (%s, %s, 'active', %s, %s)
                ON CONFLICT (user_id, product_id) DO UPDATE SET
                    status = 'active',
                    updated_at = EXCLUDED.updated_at
                """,
                (user_id, product_id, now, now),
            )


def collect_product_image_urls(rows: list[dict[str, Any]]) -> list[str]:
    values: list[str] = []
    for row in rows:
        values.extend(split_url_lines(row.get("preview_image")))
        values.extend(split_url_lines(row.get("carousel_images")))
        values.extend(split_url_lines(row.get("material_images")))
    return unique_strings([value for value in values if is_http_url(value)])


def split_url_lines(value: Any) -> list[str]:
    text = str(value or "")
    return [item.strip() for item in re.split(r"[\n\r,，;；]+", text) if item.strip()]


def variant_spec_text(row: dict[str, Any]) -> str:
    parts = []
    for index in (1, 2):
        name = clean_text(row.get(f"variant_attr_name_{index}"))
        value = clean_text(row.get(f"variant_attr_value_{index}"))
        if name and value:
            parts.append(f"{name}: {value}")
        elif value:
            parts.append(value)
    return "；".join(parts)


def variant_raw_specs(row: dict[str, Any]) -> dict[str, str]:
    specs: dict[str, str] = {}
    for index in (1, 2):
        name = clean_text(row.get(f"variant_attr_name_{index}")) or f"规格{index}"
        value = clean_text(row.get(f"variant_attr_value_{index}"))
        if value:
            specs[name] = value
    return specs


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def first_positive_number(*values: Any) -> float:
    for value in values:
        try:
            number = float(str(value).replace("¥", "").replace("$", "").replace(",", "").strip())
        except (TypeError, ValueError):
            continue
        if number > 0:
            return number
    return 0


def stable_slug(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    slug = re.sub(r"[^A-Za-z0-9_.\-\u4e00-\u9fff]+", "-", text).strip("-")
    if slug:
        return slug[:96]
    return uuid.uuid5(uuid.NAMESPACE_URL, text).hex


def product_import_id(source_product_id: str) -> str:
    return f"custom-dxm-{uuid.uuid5(uuid.NAMESPACE_URL, source_product_id).hex[:20]}"


def unique_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def is_http_url(value: Any) -> bool:
    return clean_text(value).lower().startswith(("http://", "https://"))
